import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from typing import Self


class ExcelProcessor:
    def __init__(self, input_file: str):
        self.input_file: str = input_file
        self.output_file: str = "out.xlsx"
        self.category_column: str = "ענף"
        self.value_column: str = """סכום
חיוב"""
        self.name_column: str = "שם בית עסק"
        self.df: pd.DataFrame

    def _validate_dataframe(self):
        if self.df is None:
            raise ValueError("You must run process_excel() first.")

    def process_excel(self) -> Self:

        self.df = pd.read_excel(
            self.input_file,
            skiprows=3,
            usecols=[self.category_column, self.value_column, self.name_column],
        )

        self.df[self.category_column] = self.df[self.category_column].astype(str)

        self.df = self.df.dropna(subset=[self.value_column])

        return self.sort()

    def sort(self) -> Self:
        self._validate_dataframe()

        self.df = self.df.sort_values(by=[self.category_column])
        return self

    def calculate_category_sums(self) -> pd.DataFrame:
        self._validate_dataframe()

        category_sums: pd.DataFrame = (
            self.df.groupby(self.category_column)[self.value_column].sum().reset_index()
        )
        return category_sums

    def top_5_amounts(self) -> pd.DataFrame:
        self._validate_dataframe()

        top_5 = self.df.nlargest(5, self.value_column)
        return top_5

    def style_excel_sheet(self):
        workbook = load_workbook(self.output_file)
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            for column_cells in sheet.columns:
                max_length = 0
                column = column_cells[0].column_letter
                for cell in column_cells:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except Exception:
                        pass
                adjusted_width = max_length + 2
                sheet.column_dimensions[column].width = adjusted_width

            for row in sheet:
                for cell in row:
                    cell.alignment = Alignment(horizontal="center", vertical="center")

        workbook.save(self.output_file)

    def write_to_excel(self) -> None:
        self._validate_dataframe()

        with pd.ExcelWriter(self.output_file) as writer:
            self.df.to_excel(writer, sheet_name="Processed Data", index=False)
            self.calculate_category_sums().to_excel(
                writer, sheet_name="Category Sums", index=False
            )
            self.top_5_amounts().to_excel(
                writer, sheet_name="Top 5 Amounts", index=False
            )

        self.style_excel_sheet()

    def fix_nan(self) -> Self:
        self._validate_dataframe()

        self.df.loc[
            (
                self.df[self.category_column].isna()
                | (self.df[self.category_column] == "nan")
            ),
            self.category_column,
        ] = "Unknown"

        return self

    def fix_category(self, substring: str, category: str) -> Self:
        self._validate_dataframe()

        condition = self.df[self.name_column].str.contains(
            substring, case=False, na=False
        )

        self.df.loc[condition, self.category_column] = category
        return self


if __name__ == "__main__":
    excel = (
        ExcelProcessor("data.xlsx")
        .process_excel()
        .fix_nan()
        .fix_category("google", "Subscriptions")
        .fix_category("bit", "Payments")
        .fix_category("paybox", "Payments")
        .fix_category("aliexpress", "Online")
        .fix_category("paypal", "Online")
        .sort()
        .write_to_excel()
    )
